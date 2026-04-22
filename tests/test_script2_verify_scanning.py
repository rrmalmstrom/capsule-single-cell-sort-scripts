#!/usr/bin/env python3
"""
Automated tests for verify_scanning_and_generate_ESP_files.py (Script 2)

Test strategy:
- Uses fresh_to_ESP_new_script_order as source data (post-Script-1 state)
- Creates temporary copies for tests that modify state
- Tests both happy-path and error conditions
- Critical: tests that FALSE in Checker column causes sys.exit()

Run with:
    conda run -n sip-lims python -m pytest tests/test_script2_verify_scanning.py -v
"""

import os
import sys
import shutil
import sqlite3
import tempfile
import pytest
import pandas as pd
import openpyxl
from pathlib import Path

# Add parent directory to path so we can import the script
sys.path.insert(0, str(Path(__file__).parent.parent))

import verify_scanning_and_generate_ESP_files as s2

# ---------------------------------------------------------------------------
# Paths to reference data
# ---------------------------------------------------------------------------

WORKSPACE = Path(__file__).parent.parent
# Post-Script-1 state: has library_plate_container_barcode, merged master_plate_data,
# and a completed Excel scanning file
SOURCE_PROJECT = WORKSPACE / "fresh_to_ESP_new_script_order"
GOOD_EXCEL = SOURCE_PROJECT / "4_plate_selection_and_pooling" / "B_new_plate_barcode_labels" / "599999_pool_label_scan_verificiation_tool.xlsx"
BAD_EXCEL = WORKSPACE / "fresh_to_ESP" / "4_plate_selection_and_pooling" / "C_pooling_barcode_labels" / "bad_599999_pool_label_scan_verificiation_tool.xlsx"

EXPECTED_PROPOSAL = "599999"
EXPECTED_SELECTED_PLATES = {"XUPVQ-1", "XUPVQ-3", "XUPVQ-6"}
EXPECTED_CONTAINER_BARCODES = {"27-810254", "27-000002", "27-999999"}

# Skip the entire module if the reference project directory is missing.
# The reference data is a real project snapshot that is not committed to the
# repository.  Tests will be skipped (not errored) until the data is restored.
if not SOURCE_PROJECT.exists():
    pytest.skip(
        f"Reference project directory not found: {SOURCE_PROJECT}\n"
        "Restore 'fresh_to_ESP_new_script_order/' to the workspace root to run these tests.",
        allow_module_level=True,
    )


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def tmp_project(tmp_path):
    """
    Create a fresh temporary copy of fresh_to_ESP_new_script_order for each test.
    Returns the path to the temporary project directory.
    """
    dest = tmp_path / "test_project"
    shutil.copytree(str(SOURCE_PROJECT), str(dest))
    return dest


@pytest.fixture
def db_path(tmp_project):
    """Return path to the project_summary.db in the temp project."""
    return str(tmp_project / "project_summary.db")


@pytest.fixture
def good_excel_path(tmp_project):
    """Return path to the good (all-True) Excel scanning file in the temp project."""
    return tmp_project / "4_plate_selection_and_pooling" / "B_new_plate_barcode_labels" / "599999_pool_label_scan_verificiation_tool.xlsx"


@pytest.fixture
def bad_excel_path(tmp_path):
    """
    Create a bad Excel file (with FALSE values) in a temp location.
    Uses the bad example from the workspace.
    """
    dest = tmp_path / "bad_excel"
    dest.mkdir()
    bad_dest = dest / "bad_scan.xlsx"
    shutil.copy2(str(BAD_EXCEL), str(bad_dest))
    return bad_dest


# ---------------------------------------------------------------------------
# Tests: get_proposal_from_database
# ---------------------------------------------------------------------------

class TestGetProposalFromDatabase:
    def test_returns_correct_proposal(self, tmp_project):
        proposal = s2.get_proposal_from_database(tmp_project)
        assert proposal == EXPECTED_PROPOSAL

    def test_exits_if_db_missing(self, tmp_path):
        with pytest.raises(SystemExit):
            s2.get_proposal_from_database(tmp_path / "nonexistent_dir")

    def test_exits_if_no_proposal_in_db(self, tmp_project):
        # Remove proposal data from sample_metadata
        db_path = str(tmp_project / "project_summary.db")
        conn = sqlite3.connect(db_path)
        conn.execute("UPDATE sample_metadata SET Proposal = NULL")
        conn.commit()
        conn.close()
        with pytest.raises(SystemExit):
            s2.get_proposal_from_database(tmp_project)


# ---------------------------------------------------------------------------
# Tests: find_barcode_scanning_file
# ---------------------------------------------------------------------------

class TestFindBarcodeScanningFile:
    def test_finds_existing_file(self, tmp_project):
        excel_path = s2.find_barcode_scanning_file(tmp_project, EXPECTED_PROPOSAL)
        assert excel_path.exists()
        assert excel_path.name == f"{EXPECTED_PROPOSAL}_pool_label_scan_verificiation_tool.xlsx"

    def test_exits_if_file_missing(self, tmp_project):
        with pytest.raises(SystemExit):
            s2.find_barcode_scanning_file(tmp_project, "WRONG_PROPOSAL")

    def test_exits_if_directory_missing(self, tmp_path):
        with pytest.raises(SystemExit):
            s2.find_barcode_scanning_file(tmp_path, EXPECTED_PROPOSAL)


# ---------------------------------------------------------------------------
# Tests: validate_barcode_scanning_completion
# ---------------------------------------------------------------------------

class TestValidateBarcodeScanningCompletion:
    def test_passes_with_good_excel(self, good_excel_path):
        """Good file has all True/empty in Checker column — should pass without error."""
        # Should not raise
        s2.validate_barcode_scanning_completion(good_excel_path)

    def test_exits_with_bad_excel(self, bad_excel_path):
        """Bad file has FALSE values in Checker column — must sys.exit()."""
        with pytest.raises(SystemExit):
            s2.validate_barcode_scanning_completion(bad_excel_path)

    def test_exits_if_file_missing(self, tmp_path):
        with pytest.raises(SystemExit):
            s2.validate_barcode_scanning_completion(tmp_path / "nonexistent.xlsx")

    def test_passes_with_all_empty_checker(self, tmp_path):
        """A blank template (no scans, all 'empty') should pass — no FALSE values."""
        blank = WORKSPACE / "BLANK_TEMPLATE_Pooling_Plate_Barcode_Scan_tool.xlsx"
        dest = tmp_path / "blank_copy.xlsx"
        shutil.copy2(str(blank), str(dest))
        # Blank template has no formula-evaluated values (openpyxl reads None for formulas)
        # This should pass since None is not False
        s2.validate_barcode_scanning_completion(dest)

    def test_exits_on_single_false_value(self, tmp_path):
        """Even a single FALSE in Checker column must cause sys.exit()."""
        # Create a copy of the good excel and inject one FALSE
        good_copy = tmp_path / "one_false.xlsx"
        shutil.copy2(str(GOOD_EXCEL), str(good_copy))

        # Inject FALSE into D4 (row 4, col 4) — this simulates a mismatch
        wb = openpyxl.load_workbook(good_copy)
        ws = wb['Sheet1']
        ws.cell(row=4, column=4, value=False)
        wb.save(good_copy)
        wb.close()

        with pytest.raises(SystemExit):
            s2.validate_barcode_scanning_completion(good_copy)


# ---------------------------------------------------------------------------
# Tests: read_project_database
# ---------------------------------------------------------------------------

class TestReadProjectDatabase:
    def test_reads_both_tables(self, tmp_project):
        master_df, individual_df = s2.read_project_database(tmp_project)
        assert not master_df.empty
        assert not individual_df.empty

    def test_master_has_library_plate_container_barcode(self, tmp_project):
        master_df, _ = s2.read_project_database(tmp_project)
        assert 'Library Plate Container Barcode' in master_df.columns

    def test_individual_has_library_plate_container_barcode(self, tmp_project):
        _, individual_df = s2.read_project_database(tmp_project)
        assert 'library_plate_container_barcode' in individual_df.columns

    def test_raises_if_db_missing(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            s2.read_project_database(tmp_path / "nonexistent_dir")


# ---------------------------------------------------------------------------
# Tests: archive_individual_plates_csv + regenerate_individual_plates_csv
# ---------------------------------------------------------------------------

class TestCsvArchiveAndRegenerate:
    def test_archives_existing_csv(self, tmp_project):
        csv_path = tmp_project / "individual_plates.csv"
        assert csv_path.exists(), "individual_plates.csv should exist before archiving"

        archived = s2.archive_individual_plates_csv(tmp_project)
        assert archived is not None
        assert archived.exists()
        assert "individual_plates_" in archived.name
        # Original should be gone
        assert not csv_path.exists()

    def test_archive_returns_none_if_no_csv(self, tmp_project):
        csv_path = tmp_project / "individual_plates.csv"
        if csv_path.exists():
            csv_path.unlink()
        result = s2.archive_individual_plates_csv(tmp_project)
        assert result is None

    def test_regenerate_creates_csv_with_esp_columns(self, tmp_project):
        # First update the DB with ESP status
        s2.update_individual_plates_with_esp_status(
            tmp_project, list(EXPECTED_SELECTED_PLATES), "test_batch"
        )
        # Archive old CSV
        s2.archive_individual_plates_csv(tmp_project)
        # Regenerate
        s2.regenerate_individual_plates_csv(tmp_project)

        csv_path = tmp_project / "individual_plates.csv"
        assert csv_path.exists()
        df = pd.read_csv(csv_path)
        assert 'esp_generation_status' in df.columns
        assert 'esp_generated_timestamp' in df.columns
        assert 'esp_batch_id' in df.columns


# ---------------------------------------------------------------------------
# Tests: create_smear_analysis_file
# ---------------------------------------------------------------------------

class TestCreateSmearAnalysisFile:
    def test_creates_output_directory(self, tmp_project):
        master_df, _ = s2.read_project_database(tmp_project)
        s2.create_smear_analysis_file(master_df, tmp_project)
        output_dir = tmp_project / "4_plate_selection_and_pooling" / "C_smear_file_for_ESP_upload"
        assert output_dir.exists()

    def test_creates_one_file_per_container_barcode(self, tmp_project):
        master_df, _ = s2.read_project_database(tmp_project)
        output_files = s2.create_smear_analysis_file(master_df, tmp_project)
        assert len(output_files) == 3  # 3 selected plates = 3 container barcodes

    def test_output_filenames_contain_container_barcodes(self, tmp_project):
        master_df, _ = s2.read_project_database(tmp_project)
        output_files = s2.create_smear_analysis_file(master_df, tmp_project)
        filenames = {f.name for f in output_files}
        assert "ESP_smear_file_for_upload_27-810254.csv" in filenames
        assert "ESP_smear_file_for_upload_27-000002.csv" in filenames
        assert "ESP_smear_file_for_upload_27-999999.csv" in filenames

    def test_output_files_have_correct_columns(self, tmp_project):
        master_df, _ = s2.read_project_database(tmp_project)
        output_files = s2.create_smear_analysis_file(master_df, tmp_project)
        expected_cols = [
            'Well', 'Sample ID', 'Range', 'ng/uL', '%Total', 'nmole/L',
            'Avg. Size', '%CV', 'Volume uL', 'QC Result', 'Failure Mode',
            'Index Name', 'PCR Cycles',
        ]
        for f in output_files:
            df = pd.read_csv(f)
            assert list(df.columns) == expected_cols, f"Wrong columns in {f.name}"

    def test_output_files_have_correct_fixed_values(self, tmp_project):
        master_df, _ = s2.read_project_database(tmp_project)
        output_files = s2.create_smear_analysis_file(master_df, tmp_project)
        for f in output_files:
            df = pd.read_csv(f)
            assert all(df['Range'] == '400 bp to 800 bp')
            assert all(df['%Total'] == 15)
            assert all(df['%CV'] == 20)
            assert all(df['Volume uL'] == 20)
            assert all(df['QC Result'] == 'Pass')
            assert all(df['PCR Cycles'] == 12)

    def test_exits_on_empty_dataframe(self, tmp_project):
        empty_df = pd.DataFrame()
        with pytest.raises(SystemExit):
            s2.create_smear_analysis_file(empty_df, tmp_project)

    def test_exits_if_no_nucleic_acid_id(self, tmp_project):
        master_df, _ = s2.read_project_database(tmp_project)
        # Remove all Nucleic Acid ID values
        master_df['Nucleic Acid ID'] = None
        with pytest.raises(SystemExit):
            s2.create_smear_analysis_file(master_df, tmp_project)


# ---------------------------------------------------------------------------
# Tests: update_individual_plates_with_esp_status
# ---------------------------------------------------------------------------

class TestUpdateIndividualPlatesWithEspStatus:
    def test_adds_esp_columns(self, tmp_project):
        db_path = str(tmp_project / "project_summary.db")
        s2.update_individual_plates_with_esp_status(
            tmp_project,
            list(EXPECTED_SELECTED_PLATES),
            "test_batch"
        )
        conn = sqlite3.connect(db_path)
        df = pd.read_sql("SELECT * FROM individual_plates", conn)
        conn.close()
        assert 'esp_generation_status' in df.columns
        assert 'esp_generated_timestamp' in df.columns
        assert 'esp_batch_id' in df.columns

    def test_does_not_modify_library_plate_container_barcode(self, tmp_project):
        """Script 2 must not change the container barcodes set by Script 1."""
        db_path = str(tmp_project / "project_summary.db")

        # Read before
        conn = sqlite3.connect(db_path)
        before = pd.read_sql(
            "SELECT barcode, library_plate_container_barcode FROM individual_plates",
            conn
        )
        conn.close()

        s2.update_individual_plates_with_esp_status(
            tmp_project,
            list(EXPECTED_SELECTED_PLATES),
            "test_batch"
        )

        # Read after
        conn = sqlite3.connect(db_path)
        after = pd.read_sql(
            "SELECT barcode, library_plate_container_barcode FROM individual_plates",
            conn
        )
        conn.close()

        pd.testing.assert_frame_equal(
            before.sort_values('barcode').reset_index(drop=True),
            after.sort_values('barcode').reset_index(drop=True)
        )

    def test_sets_status_to_generated_for_selected_plates(self, tmp_project):
        db_path = str(tmp_project / "project_summary.db")
        s2.update_individual_plates_with_esp_status(
            tmp_project,
            list(EXPECTED_SELECTED_PLATES),
            "test_batch"
        )
        conn = sqlite3.connect(db_path)
        df = pd.read_sql(
            "SELECT barcode, esp_generation_status FROM individual_plates WHERE selected_for_pooling=1",
            conn
        )
        conn.close()
        assert all(df['esp_generation_status'] == 'generated')

    def test_non_selected_plates_remain_pending(self, tmp_project):
        db_path = str(tmp_project / "project_summary.db")
        s2.update_individual_plates_with_esp_status(
            tmp_project,
            list(EXPECTED_SELECTED_PLATES),
            "test_batch"
        )
        conn = sqlite3.connect(db_path)
        df = pd.read_sql(
            "SELECT barcode, esp_generation_status FROM individual_plates WHERE selected_for_pooling=0",
            conn
        )
        conn.close()
        # Non-selected plates should have 'pending' (the column default)
        assert all(df['esp_generation_status'] == 'pending')

    def test_sets_batch_id(self, tmp_project):
        db_path = str(tmp_project / "project_summary.db")
        s2.update_individual_plates_with_esp_status(
            tmp_project,
            list(EXPECTED_SELECTED_PLATES),
            "my_test_batch_123"
        )
        conn = sqlite3.connect(db_path)
        df = pd.read_sql(
            "SELECT barcode, esp_batch_id FROM individual_plates WHERE selected_for_pooling=1",
            conn
        )
        conn.close()
        assert all(df['esp_batch_id'] == 'my_test_batch_123')


# ---------------------------------------------------------------------------
# Integration test: full Script 2 workflow on temp project
# ---------------------------------------------------------------------------

class TestFullScript2Workflow:
    def test_full_workflow_with_good_excel(self, tmp_project):
        """
        Run the complete Script 2 workflow with a good (all-True) Excel file.
        Verifies all expected outputs are created.
        """
        base_dir = tmp_project

        # Step 1: Get proposal
        proposal = s2.get_proposal_from_database(base_dir)
        assert proposal == EXPECTED_PROPOSAL

        # Step 2: Find Excel file
        excel_path = s2.find_barcode_scanning_file(base_dir, proposal)
        assert excel_path.exists()

        # Step 3: Validate scanning (should pass — all True)
        s2.validate_barcode_scanning_completion(excel_path)

        # Step 4: Read database
        master_df, individual_df = s2.read_project_database(base_dir)
        assert 'Library Plate Container Barcode' in master_df.columns
        assert 'library_plate_container_barcode' in individual_df.columns

        # Step 5: Generate ESP files
        output_files = s2.create_smear_analysis_file(master_df, base_dir)
        assert len(output_files) == 3

        # Step 6: Update ESP status
        grid_samples = master_df[master_df['Nucleic Acid ID'].notna()].copy()
        processed_plates = list(grid_samples['Plate_Barcode'].unique())
        batch_id = "integration_test_batch"
        s2.update_individual_plates_with_esp_status(base_dir, processed_plates, batch_id)

        # Step 7: Archive old CSV and regenerate with ESP columns
        s2.archive_individual_plates_csv(base_dir)
        s2.regenerate_individual_plates_csv(base_dir)

        # --- Assertions ---

        # Output directory exists
        output_dir = base_dir / "4_plate_selection_and_pooling" / "C_smear_file_for_ESP_upload"
        assert output_dir.exists()

        # All 3 ESP files exist with correct names
        assert (output_dir / "ESP_smear_file_for_upload_27-810254.csv").exists()
        assert (output_dir / "ESP_smear_file_for_upload_27-000002.csv").exists()
        assert (output_dir / "ESP_smear_file_for_upload_27-999999.csv").exists()

        # Each file has correct columns
        expected_cols = [
            'Well', 'Sample ID', 'Range', 'ng/uL', '%Total', 'nmole/L',
            'Avg. Size', '%CV', 'Volume uL', 'QC Result', 'Failure Mode',
            'Index Name', 'PCR Cycles',
        ]
        for f in output_files:
            df = pd.read_csv(f)
            assert list(df.columns) == expected_cols

        # Database has ESP status columns
        conn = sqlite3.connect(str(base_dir / "project_summary.db"))
        ip_df = pd.read_sql("SELECT * FROM individual_plates", conn)
        conn.close()
        assert 'esp_generation_status' in ip_df.columns
        assert 'esp_generated_timestamp' in ip_df.columns
        assert 'esp_batch_id' in ip_df.columns

        # Selected plates have 'generated' status
        selected = ip_df[ip_df['selected_for_pooling'] == 1]
        assert all(selected['esp_generation_status'] == 'generated')
        assert all(selected['esp_batch_id'] == batch_id)

        # library_plate_container_barcode is unchanged
        result = dict(zip(ip_df['barcode'], ip_df['library_plate_container_barcode']))
        assert result.get("XUPVQ-1") == "27-810254"
        assert result.get("XUPVQ-3") == "27-000002"
        assert result.get("XUPVQ-6") == "27-999999"

        # Fresh individual_plates.csv exists and has ESP columns
        csv_path = base_dir / "individual_plates.csv"
        assert csv_path.exists(), "individual_plates.csv should be regenerated"
        csv_df = pd.read_csv(csv_path)
        assert 'esp_generation_status' in csv_df.columns
        assert 'esp_generated_timestamp' in csv_df.columns
        assert 'esp_batch_id' in csv_df.columns

        print(f"\n✅ Script 2 integration test passed!")
        print(f"   Proposal: {proposal}")
        print(f"   ESP files: {[f.name for f in output_files]}")

    def test_workflow_aborts_with_bad_excel(self, tmp_project):
        """
        If the Excel file has FALSE values, the workflow must abort at validation.
        No ESP smear CSV files should be created during this run.
        """
        base_dir = tmp_project

        # Record any pre-existing ESP files before the test
        output_dir = base_dir / "4_plate_selection_and_pooling" / "C_smear_file_for_ESP_upload"
        pre_existing_files = set(output_dir.glob("*.csv")) if output_dir.exists() else set()

        # Replace the good Excel with a bad one (inject FALSE into D4)
        excel_path = base_dir / "4_plate_selection_and_pooling" / "B_new_plate_barcode_labels" / "599999_pool_label_scan_verificiation_tool.xlsx"
        wb = openpyxl.load_workbook(excel_path)
        ws = wb['Sheet1']
        ws.cell(row=4, column=4, value=False)
        wb.save(excel_path)
        wb.close()

        # Validation should fail — script must sys.exit()
        with pytest.raises(SystemExit):
            s2.validate_barcode_scanning_completion(excel_path)

        # No NEW ESP files should have been created during this aborted run
        post_files = set(output_dir.glob("*.csv")) if output_dir.exists() else set()
        new_files = post_files - pre_existing_files
        assert len(new_files) == 0, f"No new ESP files should be created after failed validation, but found: {new_files}"


# ---------------------------------------------------------------------------
# Entry point for direct execution
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
