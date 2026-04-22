#!/usr/bin/env python3
"""
Automated tests for process_grid_tables_and_generate_barcodes.py (Script 1)

Test strategy:
- Uses fresh_to_ESP_spits_1_copy as source data (never modified directly)
- Creates a temporary copy of the project directory for each test that modifies state
- Tests both happy-path and error conditions

Run with:
    conda run -n sip-lims python -m pytest tests/test_script1_process_grid_tables.py -v
"""

import os
import sys
import shutil
import sqlite3
import tempfile
import pytest
import pandas as pd
import openpyxl
from pathlib import Path

# Add parent directory to path so we can import the script
sys.path.insert(0, str(Path(__file__).parent.parent))

import process_grid_tables_and_generate_barcodes as s1

# ---------------------------------------------------------------------------
# Paths to reference data
# ---------------------------------------------------------------------------

WORKSPACE = Path(__file__).parent.parent
SOURCE_PROJECT = WORKSPACE / "fresh_to_ESP_spits_1_copy"
BLANK_TEMPLATE = WORKSPACE / "BLANK_TEMPLATE_Pooling_Plate_Barcode_Scan_tool.xlsx"

# Skip the entire module if the reference project directory is missing.
# The reference data is a real project snapshot that is not committed to the
# repository.  Tests will be skipped (not errored) until the data is restored.
if not SOURCE_PROJECT.exists():
    pytest.skip(
        f"Reference project directory not found: {SOURCE_PROJECT}\n"
        "Restore 'fresh_to_ESP_spits_1_copy/' to the workspace root to run these tests.",
        allow_module_level=True,
    )

# Expected values from the source data
EXPECTED_PROPOSAL = "599999"
EXPECTED_SELECTED_PLATES = {"XUPVQ-1", "XUPVQ-3", "XUPVQ-6"}
EXPECTED_BARCODE_MAPPING = {
    "XUPVQ-1": "27-810254",
    "XUPVQ-3": "27-000002",
    "XUPVQ-6": "27-999999",
}


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def tmp_project(tmp_path):
    """
    Create a fresh temporary copy of fresh_to_ESP_spits_1_copy for each test.
    Returns the path to the temporary project directory.
    """
    dest = tmp_path / "test_project"
    shutil.copytree(str(SOURCE_PROJECT), str(dest))
    return dest


@pytest.fixture
def db_path(tmp_project):
    """Return path to the project_summary.db in the temp project."""
    return str(tmp_project / "project_summary.db")


# ---------------------------------------------------------------------------
# Tests: read_project_database
# ---------------------------------------------------------------------------

class TestReadProjectDatabase:
    def test_reads_both_tables(self, tmp_project):
        master_df, individual_df = s1.read_project_database(tmp_project)
        assert not master_df.empty, "master_plate_data should not be empty"
        assert not individual_df.empty, "individual_plates should not be empty"

    def test_master_plate_has_required_columns(self, tmp_project):
        master_df, _ = s1.read_project_database(tmp_project)
        required = ['Plate_Barcode', 'Well', 'selected_for_pooling']
        for col in required:
            assert col in master_df.columns, f"Missing column: {col}"

    def test_individual_plates_has_required_columns(self, tmp_project):
        _, individual_df = s1.read_project_database(tmp_project)
        required = ['barcode', 'selected_for_pooling']
        for col in required:
            assert col in individual_df.columns, f"Missing column: {col}"

    def test_raises_if_db_missing(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            s1.read_project_database(tmp_path / "nonexistent_dir")


# ---------------------------------------------------------------------------
# Tests: identify_expected_grid_samples
# ---------------------------------------------------------------------------

class TestIdentifyExpectedGridSamples:
    def test_returns_correct_selected_plates(self, tmp_project):
        master_df, individual_df = s1.read_project_database(tmp_project)
        expected_samples, selected_plates = s1.identify_expected_grid_samples(
            master_df, individual_df
        )
        assert set(selected_plates) == EXPECTED_SELECTED_PLATES

    def test_expected_samples_only_from_selected_plates(self, tmp_project):
        master_df, individual_df = s1.read_project_database(tmp_project)
        expected_samples, selected_plates = s1.identify_expected_grid_samples(
            master_df, individual_df
        )
        # All expected samples must come from selected plates
        assert all(
            p in selected_plates for p in expected_samples['Plate_Barcode'].unique()
        )

    def test_expected_samples_all_selected_for_pooling(self, tmp_project):
        master_df, individual_df = s1.read_project_database(tmp_project)
        expected_samples, _ = s1.identify_expected_grid_samples(master_df, individual_df)
        assert all(expected_samples['selected_for_pooling'] == True)

    def test_exits_if_no_selected_for_pooling_column(self, tmp_project):
        master_df, individual_df = s1.read_project_database(tmp_project)
        individual_df_bad = individual_df.drop(columns=['selected_for_pooling'])
        with pytest.raises(SystemExit):
            s1.identify_expected_grid_samples(master_df, individual_df_bad)


# ---------------------------------------------------------------------------
# Tests: validate_grid_table_columns_detailed
# ---------------------------------------------------------------------------

class TestValidateGridTableColumns:
    def test_valid_grid_file_passes(self, tmp_project):
        grid_file = str(tmp_project / "4_plate_selection_and_pooling" / "grid_XUPVQ-1.csv")
        is_valid, error_msg = s1.validate_grid_table_columns_detailed(grid_file)
        assert is_valid is True
        assert error_msg is None

    def test_non_grid_csv_fails(self, tmp_project):
        non_grid = str(tmp_project / "4_plate_selection_and_pooling" / "plate_selection.csv")
        is_valid, error_msg = s1.validate_grid_table_columns_detailed(non_grid)
        assert is_valid is False
        assert "Missing columns" in error_msg

    def test_missing_file_returns_error(self, tmp_path):
        is_valid, error_msg = s1.validate_grid_table_columns_detailed(
            str(tmp_path / "nonexistent.csv")
        )
        assert is_valid is False
        assert "Error reading file" in error_msg


# ---------------------------------------------------------------------------
# Tests: find_all_grid_tables
# ---------------------------------------------------------------------------

class TestFindAllGridTables:
    def test_finds_three_grid_files(self, tmp_project):
        grid_files = s1.find_all_grid_tables(tmp_project)
        grid_names = {Path(f).name for f in grid_files}
        assert "grid_XUPVQ-1.csv" in grid_names
        assert "grid_XUPVQ-3.csv" in grid_names
        assert "grid_XUPVQ-6.csv" in grid_names

    def test_excludes_non_grid_csvs(self, tmp_project):
        grid_files = s1.find_all_grid_tables(tmp_project)
        grid_names = {Path(f).name for f in grid_files}
        assert "plate_selection.csv" not in grid_names
        assert "599999_capsule_sort_SPITS.csv" not in grid_names

    def test_exits_if_no_csv_files(self, tmp_path):
        # Create empty 4_plate_selection_and_pooling dir
        (tmp_path / "4_plate_selection_and_pooling").mkdir()
        with pytest.raises(SystemExit):
            s1.find_all_grid_tables(tmp_path)

    def test_exits_if_no_valid_grid_tables(self, tmp_path):
        # Create dir with only non-grid CSV
        pool_dir = tmp_path / "4_plate_selection_and_pooling"
        pool_dir.mkdir()
        pd.DataFrame({'col1': [1, 2]}).to_csv(pool_dir / "not_a_grid.csv", index=False)
        with pytest.raises(SystemExit):
            s1.find_all_grid_tables(tmp_path)


# ---------------------------------------------------------------------------
# Tests: read_multiple_grid_tables
# ---------------------------------------------------------------------------

class TestReadMultipleGridTables:
    def test_reads_all_three_grids(self, tmp_project):
        grid_files = s1.find_all_grid_tables(tmp_project)
        grid_dfs, combined_df = s1.read_multiple_grid_tables(grid_files)
        assert len(grid_dfs) == 3
        assert not combined_df.empty

    def test_combined_df_has_source_file_column(self, tmp_project):
        grid_files = s1.find_all_grid_tables(tmp_project)
        _, combined_df = s1.read_multiple_grid_tables(grid_files)
        assert 'Source_File' in combined_df.columns

    def test_combined_row_count_matches_sum(self, tmp_project):
        grid_files = s1.find_all_grid_tables(tmp_project)
        grid_dfs, combined_df = s1.read_multiple_grid_tables(grid_files)
        total_rows = sum(len(df) for df in grid_dfs.values())
        assert len(combined_df) == total_rows

    def test_raises_on_empty_list(self):
        with pytest.raises(ValueError):
            s1.read_multiple_grid_tables([])


# ---------------------------------------------------------------------------
# Tests: detect_duplicate_samples
# ---------------------------------------------------------------------------

class TestDetectDuplicateSamples:
    def test_no_duplicates_in_valid_data(self, tmp_project):
        grid_files = s1.find_all_grid_tables(tmp_project)
        grid_dfs, _ = s1.read_multiple_grid_tables(grid_files)
        result = s1.detect_duplicate_samples(grid_dfs)
        assert result == {}

    def test_exits_on_duplicates(self, tmp_project):
        grid_files = s1.find_all_grid_tables(tmp_project)
        grid_dfs, _ = s1.read_multiple_grid_tables(grid_files)
        # Duplicate the first grid under a different name
        first_key = list(grid_dfs.keys())[0]
        grid_dfs["duplicate_grid.csv"] = grid_dfs[first_key].copy()
        with pytest.raises(SystemExit):
            s1.detect_duplicate_samples(grid_dfs)


# ---------------------------------------------------------------------------
# Tests: validate_grid_table_completeness
# ---------------------------------------------------------------------------

class TestValidateGridTableCompleteness:
    def test_valid_data_passes(self, tmp_project):
        master_df, individual_df = s1.read_project_database(tmp_project)
        expected_samples, _ = s1.identify_expected_grid_samples(master_df, individual_df)
        grid_files = s1.find_all_grid_tables(tmp_project)
        _, combined_df = s1.read_multiple_grid_tables(grid_files)
        result = s1.validate_grid_table_completeness(expected_samples, combined_df)
        assert result is True

    def test_exits_if_sample_missing_from_grid(self, tmp_project):
        master_df, individual_df = s1.read_project_database(tmp_project)
        expected_samples, _ = s1.identify_expected_grid_samples(master_df, individual_df)
        grid_files = s1.find_all_grid_tables(tmp_project)
        _, combined_df = s1.read_multiple_grid_tables(grid_files)
        # Remove one row from combined_df to simulate missing sample
        combined_df_short = combined_df.iloc[1:].copy()
        with pytest.raises(SystemExit):
            s1.validate_grid_table_completeness(expected_samples, combined_df_short)


# ---------------------------------------------------------------------------
# Tests: extract_library_plate_container_barcode_mapping
# ---------------------------------------------------------------------------

class TestExtractBarcodeMapping:
    def test_extracts_correct_mapping(self, tmp_project):
        grid_files = s1.find_all_grid_tables(tmp_project)
        _, combined_df = s1.read_multiple_grid_tables(grid_files)
        mapping = s1.extract_library_plate_container_barcode_mapping(combined_df)
        assert mapping == EXPECTED_BARCODE_MAPPING

    def test_returns_dict(self, tmp_project):
        grid_files = s1.find_all_grid_tables(tmp_project)
        _, combined_df = s1.read_multiple_grid_tables(grid_files)
        mapping = s1.extract_library_plate_container_barcode_mapping(combined_df)
        assert isinstance(mapping, dict)


# ---------------------------------------------------------------------------
# Tests: validate_and_merge_data
# ---------------------------------------------------------------------------

class TestValidateAndMergeData:
    def test_merge_produces_correct_row_count(self, tmp_project):
        master_df, individual_df = s1.read_project_database(tmp_project)
        expected_samples, _ = s1.identify_expected_grid_samples(master_df, individual_df)
        grid_files = s1.find_all_grid_tables(tmp_project)
        _, combined_df = s1.read_multiple_grid_tables(grid_files)
        merged_df = s1.validate_and_merge_data(master_df, expected_samples, combined_df)
        # Merged should have same rows as master (left join)
        assert len(merged_df) == len(master_df)

    def test_merged_has_nucleic_acid_id_for_selected(self, tmp_project):
        master_df, individual_df = s1.read_project_database(tmp_project)
        expected_samples, selected_plates = s1.identify_expected_grid_samples(
            master_df, individual_df
        )
        grid_files = s1.find_all_grid_tables(tmp_project)
        _, combined_df = s1.read_multiple_grid_tables(grid_files)
        merged_df = s1.validate_and_merge_data(master_df, expected_samples, combined_df)
        # Only rows that are selected_for_pooling=True should have Nucleic Acid ID
        # (not all rows from selected plates — only the pooling-selected wells)
        selected_rows = merged_df[merged_df['selected_for_pooling'] == True]
        assert selected_rows['Nucleic Acid ID'].notna().all(), (
            "All selected_for_pooling rows should have a Nucleic Acid ID after merge"
        )

    def test_merged_has_container_barcode_column(self, tmp_project):
        master_df, individual_df = s1.read_project_database(tmp_project)
        expected_samples, _ = s1.identify_expected_grid_samples(master_df, individual_df)
        grid_files = s1.find_all_grid_tables(tmp_project)
        _, combined_df = s1.read_multiple_grid_tables(grid_files)
        merged_df = s1.validate_and_merge_data(master_df, expected_samples, combined_df)
        assert 'Library Plate Container Barcode' in merged_df.columns


# ---------------------------------------------------------------------------
# Tests: get_proposal_value
# ---------------------------------------------------------------------------

class TestGetProposalValue:
    def test_returns_correct_proposal(self, db_path):
        proposal = s1.get_proposal_value(db_path)
        assert proposal == EXPECTED_PROPOSAL

    def test_raises_if_db_missing(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            s1.get_proposal_value(str(tmp_path / "nonexistent.db"))


# ---------------------------------------------------------------------------
# Tests: populate_excel_template
# ---------------------------------------------------------------------------

class TestPopulateExcelTemplate:
    def test_populates_column_c_correctly(self, tmp_path):
        # Copy blank template to temp dir
        dest = str(tmp_path / "test_template.xlsx")
        shutil.copy2(str(BLANK_TEMPLATE), dest)

        plate_data = [
            ("XUPVQ-1", "27-810254"),
            ("XUPVQ-3", "27-000002"),
            ("XUPVQ-6", "27-999999"),
        ]
        s1.populate_excel_template(dest, plate_data)

        # Verify with openpyxl
        wb = openpyxl.load_workbook(dest)
        ws = wb['Sheet1']

        # Plate 1: rows 2 (plate name) and 3 (LIMS ID)
        assert ws.cell(row=2, column=3).value == "XUPVQ-1"
        assert ws.cell(row=3, column=3).value == "27-810254"
        # Plate 2: rows 4 and 5
        assert ws.cell(row=4, column=3).value == "XUPVQ-3"
        assert ws.cell(row=5, column=3).value == "27-000002"
        # Plate 3: rows 6 and 7
        assert ws.cell(row=6, column=3).value == "XUPVQ-6"
        assert ws.cell(row=7, column=3).value == "27-999999"
        wb.close()

    def test_does_not_modify_other_columns(self, tmp_path):
        dest = str(tmp_path / "test_template.xlsx")
        shutil.copy2(str(BLANK_TEMPLATE), dest)

        # Read original column A and B values before populating
        wb_before = openpyxl.load_workbook(dest)
        ws_before = wb_before['Sheet1']
        col_a_before = [ws_before.cell(row=r, column=1).value for r in range(1, 10)]
        col_b_before = [ws_before.cell(row=r, column=2).value for r in range(1, 10)]
        wb_before.close()

        plate_data = [("XUPVQ-1", "27-810254")]
        s1.populate_excel_template(dest, plate_data)

        wb_after = openpyxl.load_workbook(dest)
        ws_after = wb_after['Sheet1']
        col_a_after = [ws_after.cell(row=r, column=1).value for r in range(1, 10)]
        col_b_after = [ws_after.cell(row=r, column=2).value for r in range(1, 10)]
        wb_after.close()

        assert col_a_before == col_a_after, "Column A should not be modified"
        assert col_b_before == col_b_after, "Column B should not be modified"

    def test_empty_rows_remain_empty_after_3_plates(self, tmp_path):
        dest = str(tmp_path / "test_template.xlsx")
        shutil.copy2(str(BLANK_TEMPLATE), dest)

        plate_data = [
            ("XUPVQ-1", "27-810254"),
            ("XUPVQ-3", "27-000002"),
            ("XUPVQ-6", "27-999999"),
        ]
        s1.populate_excel_template(dest, plate_data)

        wb = openpyxl.load_workbook(dest)
        ws = wb['Sheet1']
        # Row 8 (plate 4 "plate name") should still be empty in column C
        assert ws.cell(row=8, column=3).value is None
        assert ws.cell(row=9, column=3).value is None
        wb.close()

    def test_raises_if_too_many_plates(self, tmp_path):
        dest = str(tmp_path / "test_template.xlsx")
        shutil.copy2(str(BLANK_TEMPLATE), dest)
        plate_data = [("PLATE-%d" % i, "27-%06d" % i) for i in range(21)]
        with pytest.raises(ValueError, match="Too many plates"):
            s1.populate_excel_template(dest, plate_data)


# ---------------------------------------------------------------------------
# Tests: create_bartender_file
# ---------------------------------------------------------------------------

class TestCreateBartenderFile:
    def test_creates_file(self, tmp_path):
        plate_data = [("XUPVQ-1", "27-810254"), ("XUPVQ-3", "27-000002")]
        output_dir = str(tmp_path / "output")
        path = s1.create_bartender_file(plate_data, output_dir, "599999")
        assert Path(path).exists()

    def test_filename_contains_proposal(self, tmp_path):
        plate_data = [("XUPVQ-1", "27-810254")]
        output_dir = str(tmp_path / "output")
        path = s1.create_bartender_file(plate_data, output_dir, "599999")
        assert "599999" in Path(path).name

    def test_file_contains_container_barcodes(self, tmp_path):
        plate_data = [("XUPVQ-1", "27-810254"), ("XUPVQ-3", "27-000002")]
        output_dir = str(tmp_path / "output")
        path = s1.create_bartender_file(plate_data, output_dir, "599999")
        content = Path(path).read_text()
        assert "27-810254" in content
        assert "27-000002" in content

    def test_file_has_bartender_header(self, tmp_path):
        plate_data = [("XUPVQ-1", "27-810254")]
        output_dir = str(tmp_path / "output")
        path = s1.create_bartender_file(plate_data, output_dir, "599999")
        content = Path(path).read_text()
        assert "%BTW%" in content
        assert "%END%" in content


# ---------------------------------------------------------------------------
# Tests: copy_template_file
# ---------------------------------------------------------------------------

class TestCopyTemplateFile:
    def test_creates_file_with_proposal_name(self, tmp_path):
        output_dir = str(tmp_path / "output")
        path = s1.copy_template_file(str(BLANK_TEMPLATE), output_dir, "599999")
        assert Path(path).exists()
        assert "599999" in Path(path).name
        assert Path(path).name == "599999_pool_label_scan_verificiation_tool.xlsx"

    def test_raises_if_template_missing(self, tmp_path):
        with pytest.raises(FileNotFoundError):
            s1.copy_template_file(
                str(tmp_path / "nonexistent.xlsx"),
                str(tmp_path / "output"),
                "599999"
            )


# ---------------------------------------------------------------------------
# Tests: get_pooling_plates_data (after database update)
# ---------------------------------------------------------------------------

class TestGetPoolingPlatesData:
    def test_exits_if_column_missing(self, db_path):
        """Before Script 1 runs, library_plate_container_barcode column doesn't exist → sys.exit()."""
        with pytest.raises(SystemExit):
            s1.get_pooling_plates_data(db_path)

    def test_returns_data_after_update(self, tmp_project):
        """After update_individual_plates_with_container_barcodes, data should be available."""
        db_path = str(tmp_project / "project_summary.db")
        barcode_mapping = EXPECTED_BARCODE_MAPPING
        processed_plates = list(EXPECTED_SELECTED_PLATES)

        s1.update_individual_plates_with_container_barcodes(
            tmp_project, processed_plates, barcode_mapping
        )

        result = s1.get_pooling_plates_data(db_path)
        assert len(result) == 3
        result_dict = dict(result)
        assert result_dict.get("XUPVQ-1") == "27-810254"
        assert result_dict.get("XUPVQ-3") == "27-000002"
        assert result_dict.get("XUPVQ-6") == "27-999999"


# ---------------------------------------------------------------------------
# Tests: update_individual_plates_with_container_barcodes
# ---------------------------------------------------------------------------

class TestUpdateIndividualPlates:
    def test_adds_container_barcode_column(self, tmp_project):
        db_path = str(tmp_project / "project_summary.db")
        s1.update_individual_plates_with_container_barcodes(
            tmp_project,
            list(EXPECTED_SELECTED_PLATES),
            EXPECTED_BARCODE_MAPPING
        )
        conn = sqlite3.connect(db_path)
        df = pd.read_sql("SELECT * FROM individual_plates", conn)
        conn.close()
        assert 'library_plate_container_barcode' in df.columns
        # ESP columns should NOT be added by Script 1
        assert 'esp_generation_status' not in df.columns
        assert 'esp_generated_timestamp' not in df.columns
        assert 'esp_batch_id' not in df.columns

    def test_sets_container_barcodes_correctly(self, tmp_project):
        db_path = str(tmp_project / "project_summary.db")
        s1.update_individual_plates_with_container_barcodes(
            tmp_project,
            list(EXPECTED_SELECTED_PLATES),
            EXPECTED_BARCODE_MAPPING
        )
        conn = sqlite3.connect(db_path)
        df = pd.read_sql("SELECT barcode, library_plate_container_barcode FROM individual_plates", conn)
        conn.close()
        result = dict(zip(df['barcode'], df['library_plate_container_barcode']))
        assert result.get("XUPVQ-1") == "27-810254"
        assert result.get("XUPVQ-3") == "27-000002"
        assert result.get("XUPVQ-6") == "27-999999"
        # Non-selected plates should have NULL
        assert result.get("XUPVQ-2") is None

    def test_exits_if_barcode_missing_from_mapping(self, tmp_project):
        """If a selected plate has no container barcode in the mapping, script must exit."""
        incomplete_mapping = {"XUPVQ-1": "27-810254"}  # Missing XUPVQ-3 and XUPVQ-6
        with pytest.raises(SystemExit):
            s1.update_individual_plates_with_container_barcodes(
                tmp_project,
                list(EXPECTED_SELECTED_PLATES),
                incomplete_mapping
            )


# ---------------------------------------------------------------------------
# Tests: archive_grid_table_files
# ---------------------------------------------------------------------------

class TestArchiveGridTableFiles:
    def test_moves_files_to_archive_dir(self, tmp_project):
        grid_files = s1.find_all_grid_tables(tmp_project)
        s1.archive_grid_table_files(tmp_project, grid_files)

        archive_dir = tmp_project / "4_plate_selection_and_pooling" / "previously_processed_grid_files"
        assert archive_dir.exists()
        archived = list(archive_dir.glob("*.csv"))
        assert len(archived) == 3

    def test_original_files_no_longer_in_place(self, tmp_project):
        grid_files = s1.find_all_grid_tables(tmp_project)
        s1.archive_grid_table_files(tmp_project, grid_files)

        pool_dir = tmp_project / "4_plate_selection_and_pooling"
        remaining_grids = [
            f for f in pool_dir.glob("*.csv")
            if f.name.startswith("grid_")
        ]
        assert len(remaining_grids) == 0


# ---------------------------------------------------------------------------
# Integration test: full Script 1 workflow on temp project
# ---------------------------------------------------------------------------

class TestFullScript1Workflow:
    def test_full_workflow_produces_expected_outputs(self, tmp_project):
        """
        Run the complete Script 1 workflow and verify all expected outputs exist.
        """
        # Run the full workflow by calling main functions in sequence
        base_dir = tmp_project
        script_dir = WORKSPACE  # template is in workspace root

        # Step 1: Read database
        master_df, individual_df = s1.read_project_database(base_dir)

        # Step 2: Identify expected samples
        expected_samples, selected_plates = s1.identify_expected_grid_samples(
            master_df, individual_df
        )

        # Step 3: Find grid tables
        grid_files = s1.find_all_grid_tables(base_dir)

        # Step 4: Read and validate
        grid_dfs, combined_df = s1.read_multiple_grid_tables(grid_files)
        s1.detect_duplicate_samples(grid_dfs)
        s1.validate_grid_table_completeness(expected_samples, combined_df)

        # Step 5: Extract barcode mapping
        barcode_mapping = s1.extract_library_plate_container_barcode_mapping(combined_df)

        # Step 6: Merge
        merged_df = s1.validate_and_merge_data(master_df, expected_samples, combined_df)

        # Step 7: Update database
        processed_plates = list(
            merged_df[merged_df['Nucleic Acid ID'].notna()]['Plate_Barcode'].unique()
        )
        s1.update_individual_plates_with_container_barcodes(
            base_dir, processed_plates, barcode_mapping
        )
        s1.update_master_plate_data_table(base_dir, merged_df)

        # Step 8: Archive and regenerate CSVs
        s1.archive_grid_table_files(base_dir, grid_files)
        s1.archive_csv_files(base_dir)
        s1.generate_fresh_csv_files(base_dir, merged_df)

        # Step 9: Generate barcode materials
        db_path = str(base_dir / "project_summary.db")
        proposal = s1.get_proposal_value(db_path)
        plate_data = s1.get_pooling_plates_data(db_path)

        output_dir = str(base_dir / "4_plate_selection_and_pooling" / "B_new_plate_barcode_labels")
        template_path = s1.find_blank_template(script_dir)
        excel_path = s1.copy_template_file(str(template_path), output_dir, proposal)
        s1.populate_excel_template(excel_path, plate_data)
        bartender_path = s1.create_bartender_file(plate_data, output_dir, proposal)

        # --- Assertions ---

        # Excel file exists with correct name
        assert Path(excel_path).exists()
        assert Path(excel_path).name == f"{EXPECTED_PROPOSAL}_pool_label_scan_verificiation_tool.xlsx"

        # BarTender file exists
        assert Path(bartender_path).exists()

        # Excel has correct barcode data in column C
        wb = openpyxl.load_workbook(excel_path)
        ws = wb['Sheet1']
        # plate_data is sorted by barcode: XUPVQ-1, XUPVQ-3, XUPVQ-6
        plate_data_dict = dict(plate_data)
        for i, (plate_barcode, container_barcode) in enumerate(plate_data):
            plate_name_row = 2 + (i * 2)
            lims_id_row = 3 + (i * 2)
            assert ws.cell(row=plate_name_row, column=3).value == plate_barcode
            assert ws.cell(row=lims_id_row, column=3).value == container_barcode
        wb.close()

        # Grid files archived
        archive_dir = base_dir / "4_plate_selection_and_pooling" / "previously_processed_grid_files"
        assert archive_dir.exists()
        assert len(list(archive_dir.glob("grid_*.csv"))) == 3

        # Fresh CSVs generated
        assert (base_dir / "master_plate_data.csv").exists()
        assert (base_dir / "individual_plates.csv").exists()

        # Database has container barcodes
        conn = sqlite3.connect(str(base_dir / "project_summary.db"))
        df = pd.read_sql("SELECT barcode, library_plate_container_barcode FROM individual_plates", conn)
        conn.close()
        result = dict(zip(df['barcode'], df['library_plate_container_barcode']))
        assert result.get("XUPVQ-1") == "27-810254"
        assert result.get("XUPVQ-3") == "27-000002"
        assert result.get("XUPVQ-6") == "27-999999"

        print(f"\n✅ Integration test passed!")
        print(f"   Proposal: {proposal}")
        print(f"   Plates processed: {processed_plates}")
        print(f"   Excel: {Path(excel_path).name}")
        print(f"   BarTender: {Path(bartender_path).name}")


# ---------------------------------------------------------------------------
# Entry point for direct execution
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    pytest.main([__file__, "-v", "--tb=short"])
