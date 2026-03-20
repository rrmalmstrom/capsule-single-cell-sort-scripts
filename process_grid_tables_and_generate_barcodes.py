#!/usr/bin/env python3
"""
Grid Table Processing and Barcode Generation Script

New Script 1 in the refactored ESP workflow.

Workflow position:
  [This script] → Manual barcode scanning (Excel) → verify_scanning_and_generate_ESP_files.py

This script:
1. Reads the project database (project_summary.db)
2. Identifies samples selected for pooling
3. Finds and validates grid table CSV files in 4_plate_selection_and_pooling/
4. Reads and combines grid tables, validates completeness
5. Extracts Library Plate Container Barcode mappings
6. Updates the database (individual_plates + master_plate_data tables)
7. Archives processed grid files and regenerates CSV files
8. Generates BarTender label file for container barcodes
9. Copies and populates the Excel barcode scanning template (Column C only)

Output directory: 4_plate_selection_and_pooling/B_new_plate_barcode_labels/

Key differences from make_ESP_smear_analysis_file.py:
- Does NOT generate ESP smear files (that is Script 2's job)
- DOES generate barcode scanning materials (Excel template + BarTender file)
"""

import os
import sys
import shutil
import sqlite3
import openpyxl
import pandas as pd
import argparse
from pathlib import Path
from datetime import datetime
from typing import List, Tuple, Optional


# ---------------------------------------------------------------------------
# Success marker
# ---------------------------------------------------------------------------

def create_success_marker():
    """Create success marker file for workflow manager integration."""
    script_name = Path(__file__).stem
    status_dir = Path(".workflow_status")
    status_dir.mkdir(exist_ok=True)
    success_file = status_dir / f"{script_name}.success"

    try:
        with open(success_file, "w") as f:
            f.write(f"SUCCESS: {script_name} completed at {datetime.now()}\n")
        print(f"✅ Success marker created: {success_file}")
    except Exception as e:
        print(f"FATAL ERROR: Could not create success marker: {e}")
        print("Laboratory automation requires workflow integration for safety.")
        sys.exit()


# ---------------------------------------------------------------------------
# Database reading
# ---------------------------------------------------------------------------

def read_project_database(base_dir):
    """Read the ESP project database from project_summary.db into pandas DataFrames."""
    db_path = Path(base_dir) / "project_summary.db"

    if not db_path.exists():
        raise FileNotFoundError(f"Database not found: {db_path}")

    try:
        conn = sqlite3.connect(db_path)
        master_plate_df = pd.read_sql_query("SELECT * FROM master_plate_data", conn)
        individual_plates_df = pd.read_sql_query("SELECT * FROM individual_plates", conn)
        conn.close()
        return master_plate_df, individual_plates_df

    except Exception as e:
        print(f"ERROR reading database: {e}")
        raise


# ---------------------------------------------------------------------------
# Grid table discovery and validation
# ---------------------------------------------------------------------------

def identify_expected_grid_samples(master_plate_df, individual_plates_df):
    """
    Identify which samples should be present in grid tables based on pooling selection.

    Logic:
    1. Find plates selected for pooling from individual_plates table
    2. Find samples selected for pooling from master_plate_data table
    3. Return samples that meet both criteria
    """
    if 'selected_for_pooling' not in individual_plates_df.columns:
        print("ERROR: Missing 'selected_for_pooling' column in individual_plates table")
        print("SCRIPT TERMINATED: Cannot determine which plates are selected for pooling")
        sys.exit()

    if 'selected_for_pooling' not in master_plate_df.columns:
        print("ERROR: Missing 'selected_for_pooling' column in master_plate_data table")
        print("SCRIPT TERMINATED: Cannot determine which samples are selected for pooling")
        sys.exit()

    selected_plates = individual_plates_df[
        individual_plates_df['selected_for_pooling'] == True
    ]['barcode'].tolist()

    expected_samples = master_plate_df[
        (master_plate_df['selected_for_pooling'] == True) &
        (master_plate_df['Plate_Barcode'].isin(selected_plates))
    ].copy()

    if len(expected_samples) == 0:
        print("ERROR: No samples found that are selected for pooling from plates selected for pooling")
        print("SCRIPT TERMINATED: No expected grid table samples identified")
        sys.exit()

    return expected_samples, selected_plates


def validate_grid_table_columns_detailed(csv_file):
    """Check if CSV file has required grid table columns with detailed error reporting."""
    required_cols = [
        'Well',
        'Library Plate Label',
        'Illumina Library',
        'Library Plate Container Barcode',
        'Nucleic Acid ID',
    ]

    try:
        df_header = pd.read_csv(csv_file, nrows=0)
        missing_cols = [col for col in required_cols if col not in df_header.columns]

        if not missing_cols:
            return True, None
        else:
            return False, f"Missing columns: {missing_cols}"

    except Exception as e:
        return False, f"Error reading file: {e}"


def find_csv_files(base_dir):
    """Find all CSV files in the 4_plate_selection_and_pooling subdirectory (top-level only)."""
    csv_files = []
    try:
        grid_table_dir = Path(base_dir) / "4_plate_selection_and_pooling"
        if not grid_table_dir.exists():
            print(f"ERROR: Grid table directory not found: {grid_table_dir}")
            print("SCRIPT TERMINATED: Required grid table directory does not exist")
            sys.exit()

        for file_path in grid_table_dir.glob("*.csv"):
            csv_files.append(str(file_path))
        return csv_files
    except Exception as e:
        print(f"ERROR finding CSV files: {e}")
        return []


def find_all_grid_tables(base_dir):
    """
    Find and validate ALL grid table files in the 4_plate_selection_and_pooling directory.

    Scans for CSV files and validates each one for required grid table columns.
    Returns all valid files for multi-grid processing.

    Required Columns:
        - Well
        - Library Plate Label
        - Illumina Library
        - Library Plate Container Barcode
        - Nucleic Acid ID
    """
    csv_files = find_csv_files(base_dir)

    if not csv_files:
        print("ERROR: No CSV files found in 4_plate_selection_and_pooling directory")
        print("SCRIPT TERMINATED: Grid table directory contains no CSV files")
        sys.exit()

    valid_files = []
    invalid_files = []

    for csv_file in csv_files:
        is_valid, error_msg = validate_grid_table_columns_detailed(csv_file)
        if is_valid:
            valid_files.append(csv_file)
        else:
            invalid_files.append((csv_file, error_msg))

    if len(valid_files) == 0:
        print(f"ERROR: No valid grid table found in 4_plate_selection_and_pooling directory")
        print(f"Found {len(csv_files)} CSV file(s), but none contain the required columns:")
        for csv_file, error_msg in invalid_files:
            print(f"  - {Path(csv_file).name}: {error_msg}")
        print("A grid table CSV file must contain: Well, Library Plate Label, Illumina Library, "
              "Library Plate Container Barcode, Nucleic Acid ID")
        print("SCRIPT TERMINATED: No valid grid table files found")
        sys.exit()

    return valid_files


# ---------------------------------------------------------------------------
# Grid table reading and validation
# ---------------------------------------------------------------------------

def read_multiple_grid_tables(grid_table_files):
    """Read and combine multiple grid table files."""
    if not grid_table_files:
        raise ValueError("No grid table files provided")

    grid_dataframes = {}
    combined_grid_df = pd.DataFrame()

    for filename in grid_table_files:
        grid_path = Path(filename)

        try:
            grid_df = pd.read_csv(filename)
            grid_dataframes[grid_path.name] = grid_df
            grid_df['Source_File'] = grid_path.name

            if combined_grid_df.empty:
                combined_grid_df = grid_df.copy()
            else:
                combined_grid_df = pd.concat([combined_grid_df, grid_df], ignore_index=True)

        except Exception as e:
            print(f"ERROR processing {filename}: {e}")
            continue

    if combined_grid_df.empty:
        raise ValueError("No valid grid table data could be read")

    return grid_dataframes, combined_grid_df


def detect_duplicate_samples(grid_dataframes):
    """Detect duplicate samples across grid tables."""
    all_samples = []
    sample_sources = {}

    for filename, df in grid_dataframes.items():
        for _, row in df.iterrows():
            sample_id = row.get('Nucleic Acid ID', '')
            if pd.notna(sample_id) and sample_id != '':
                sample_key = f"{sample_id}_{row.get('Well', '')}_{row.get('Library Plate Label', '')}"
                all_samples.append(sample_key)

                if sample_key in sample_sources:
                    sample_sources[sample_key].append(filename)
                else:
                    sample_sources[sample_key] = [filename]

    duplicates = {sample: sources for sample, sources in sample_sources.items()
                  if len(sources) > 1}

    if duplicates:
        print(f"ERROR: Found {len(duplicates)} duplicate samples:")
        for sample, sources in duplicates.items():
            print(f"  {sample} appears in: {', '.join(sources)}")
        print("SCRIPT TERMINATED: Duplicate samples detected - data integrity compromised")
        sys.exit()
    else:
        return {}


def validate_grid_table_completeness(expected_samples, combined_grid_df):
    """
    Validate that grid table contains exactly the expected samples (no more, no less).
    """
    expected_set = set(zip(expected_samples['Plate_Barcode'], expected_samples['Well']))
    grid_set = set(zip(combined_grid_df['Library Plate Label'], combined_grid_df['Well']))

    missing_from_grid = expected_set - grid_set
    unexpected_in_grid = grid_set - expected_set

    if missing_from_grid:
        print(f"ERROR: Found {len(missing_from_grid)} expected samples missing from grid tables:")
        for plate_barcode, well in missing_from_grid:
            print(f"  {plate_barcode} - {well}")
        print("SCRIPT TERMINATED: Expected samples missing from grid tables")
        sys.exit()

    if unexpected_in_grid:
        print(f"ERROR: Found {len(unexpected_in_grid)} unexpected samples in grid tables:")
        for plate_barcode, well in unexpected_in_grid:
            print(f"  {plate_barcode} - {well}")
        print("SCRIPT TERMINATED: Grid tables contain samples not selected for pooling")
        sys.exit()

    return True


def validate_and_merge_data(master_plate_df, expected_samples, grid_df):
    """
    Validate and merge grid table data with master plate dataframe.
    Only expected samples should have grid data, but all master data is preserved.

    ESP merging strategy:
    - Grid table: ['Library Plate Label', 'Well']
    - Master plate data: ['Plate_Barcode', 'Well']
    """
    grid_merge_cols = ['Library Plate Label', 'Well']
    db_merge_cols = ['Plate_Barcode', 'Well']

    missing_grid_cols = [col for col in grid_merge_cols if col not in grid_df.columns]
    missing_db_cols = [col for col in db_merge_cols if col not in master_plate_df.columns]

    if missing_grid_cols:
        print(f"ERROR: Missing grid table columns: {missing_grid_cols}")
        print("SCRIPT TERMINATED: Required columns missing from grid table")
        sys.exit()
    if missing_db_cols:
        print(f"ERROR: Missing master plate data columns: {missing_db_cols}")
        print("SCRIPT TERMINATED: Required columns missing from database")
        sys.exit()

    grid_data_cols = ['Illumina Library', 'Nucleic Acid ID', 'Library Plate Container Barcode']
    grid_merge_df = grid_df[grid_merge_cols + grid_data_cols].copy()

    grid_merge_df = grid_merge_df.rename(columns={
        'Library Plate Label': 'Plate_Barcode'
    })

    merged_df = pd.merge(
        master_plate_df,
        grid_merge_df,
        on=['Plate_Barcode', 'Well'],
        how='left',
        suffixes=(None, '_y')
    )

    columns_to_remove = ['Well_y']
    for col in columns_to_remove:
        if col in merged_df.columns:
            merged_df = merged_df.drop(columns=[col])

    missing_grid_data = []
    for _, row in expected_samples.iterrows():
        merged_row = merged_df[
            (merged_df['Plate_Barcode'] == row['Plate_Barcode']) &
            (merged_df['Well'] == row['Well'])
        ]
        if merged_row.empty or pd.isna(merged_row.iloc[0]['Nucleic Acid ID']):
            missing_grid_data.append((row['Plate_Barcode'], row['Well']))

    if missing_grid_data:
        print("ERROR: Some expected samples are missing grid table data:")
        for plate_barcode, well in missing_grid_data:
            print(f"  {plate_barcode} - {well}")
        print("SCRIPT TERMINATED: Incomplete merge - not all expected samples have grid data")
        sys.exit()

    return merged_df


# ---------------------------------------------------------------------------
# Barcode mapping extraction
# ---------------------------------------------------------------------------

def extract_library_plate_container_barcode_mapping(combined_grid_df):
    """
    Extract mapping from Library Plate Label to Library Plate Container Barcode.

    Returns:
        dict: e.g. {'XUPVQ-1': '27-810254', 'XUPVQ-3': '27-000002'}
    """
    mapping_df = combined_grid_df[
        ['Library Plate Label', 'Library Plate Container Barcode']
    ].drop_duplicates()
    barcode_mapping = dict(
        zip(mapping_df['Library Plate Label'], mapping_df['Library Plate Container Barcode'])
    )
    return barcode_mapping


# ---------------------------------------------------------------------------
# Database update functions
# ---------------------------------------------------------------------------

def update_individual_plates_with_container_barcodes(base_dir, processed_plate_barcodes, barcode_mapping):
    """
    Update individual_plates table with Library Plate Container Barcodes from grid tables.

    Script 1 only adds library_plate_container_barcode — no ESP-related columns.
    ESP status columns (esp_generation_status, esp_generated_timestamp, esp_batch_id)
    are added by Script 2 when ESP smear files are actually generated.

    Adds column if it doesn't exist:
        - library_plate_container_barcode (TEXT)
    """
    if not processed_plate_barcodes:
        return

    sql_db_path = Path(base_dir) / 'project_summary.db'

    try:
        from sqlalchemy import create_engine, text
        engine = create_engine(f'sqlite:///{sql_db_path}')

        with engine.connect() as conn:
            result = conn.execute(text("PRAGMA table_info(individual_plates)"))
            existing_columns = [row[1] for row in result]

            if 'library_plate_container_barcode' not in existing_columns:
                conn.execute(text(
                    "ALTER TABLE individual_plates ADD COLUMN library_plate_container_barcode TEXT"
                ))
                conn.commit()

            for barcode in processed_plate_barcodes:
                container_barcode = barcode_mapping.get(barcode)

                if not container_barcode:
                    print(f"ERROR: No Library Plate Container Barcode found for plate {barcode}")
                    print("SCRIPT TERMINATED: All selected plates must have a container barcode in the grid table")
                    sys.exit()

                update_query = text("""
                    UPDATE individual_plates
                    SET library_plate_container_barcode = :container_barcode
                    WHERE barcode = :barcode
                """)

                conn.execute(update_query, {
                    'barcode': barcode,
                    'container_barcode': container_barcode,
                })

            conn.commit()

    except Exception as e:
        print(f"ERROR updating individual_plates table: {e}")
        raise
    finally:
        engine.dispose()


def update_master_plate_data_table(base_dir, merged_df):
    """
    Replace master_plate_data table with the merged dataframe.
    """
    sql_db_path = Path(base_dir) / 'project_summary.db'

    try:
        from sqlalchemy import create_engine
        engine = create_engine(f'sqlite:///{sql_db_path}')
        merged_df.to_sql('master_plate_data', engine, if_exists='replace', index=False)

    except Exception as e:
        print(f"ERROR updating master_plate_data table: {e}")
        raise
    finally:
        engine.dispose()


# ---------------------------------------------------------------------------
# File archiving
# ---------------------------------------------------------------------------

def archive_grid_table_files(base_dir, grid_table_files):
    """
    Move processed grid table CSV files to previously_processed_grid_files/ subfolder.
    """
    archive_dir = (
        Path(base_dir) / "4_plate_selection_and_pooling" / "previously_processed_grid_files"
    )
    archive_dir.mkdir(parents=True, exist_ok=True)

    moved_files = []
    for grid_file_path in grid_table_files:
        grid_file = Path(grid_file_path)
        if grid_file.exists():
            dest_path = archive_dir / grid_file.name
            grid_file.rename(dest_path)
            moved_files.append(dest_path)
        else:
            print(f"⚠️  Grid table file not found for archiving: {grid_file_path}")

    return moved_files


def archive_database_file(base_dir):
    """
    Archive existing database file with timestamp suffix by copying (not moving).
    """
    db_path = Path(base_dir) / "project_summary.db"

    if db_path.exists():
        timestamp = datetime.now().strftime("%Y_%m_%d-Time%H-%M-%S")
        archive_dir = Path(base_dir) / "archived_files"
        archive_dir.mkdir(exist_ok=True)

        archive_name = f"{db_path.stem}_{timestamp}{db_path.suffix}"
        archive_path = archive_dir / archive_name

        shutil.copy2(str(db_path), str(archive_path))
        return archive_path
    return None


def archive_csv_files(base_dir):
    """
    Archive existing master_plate_data.csv and individual_plates.csv with timestamp suffix.
    """
    timestamp = datetime.now().strftime("%Y_%m_%d-Time%H-%M-%S")
    archive_dir = Path(base_dir) / "archived_files"
    archive_dir.mkdir(exist_ok=True)

    archived_files = {}

    master_csv_path = Path(base_dir) / "master_plate_data.csv"
    if master_csv_path.exists():
        archive_name = f"master_plate_data_{timestamp}.csv"
        archive_path = archive_dir / archive_name
        shutil.move(str(master_csv_path), str(archive_path))
        archived_files['master_plate_data'] = archive_path

    plates_csv_path = Path(base_dir) / "individual_plates.csv"
    if plates_csv_path.exists():
        archive_name = f"individual_plates_{timestamp}.csv"
        archive_path = archive_dir / archive_name
        shutil.move(str(plates_csv_path), str(archive_path))
        archived_files['individual_plates'] = archive_path

    return archived_files


def generate_fresh_csv_files(base_dir, updated_master_df):
    """
    Generate fresh CSV files from updated database tables.
    """
    try:
        master_csv_path = Path(base_dir) / "master_plate_data.csv"
        updated_master_df.to_csv(master_csv_path, index=False)

        sql_db_path = Path(base_dir) / 'project_summary.db'
        from sqlalchemy import create_engine
        engine = create_engine(f'sqlite:///{sql_db_path}')
        individual_plates_df = pd.read_sql('SELECT * FROM individual_plates', engine)
        engine.dispose()

        plates_csv_path = Path(base_dir) / "individual_plates.csv"
        individual_plates_df.to_csv(plates_csv_path, index=False)

    except Exception as e:
        print(f"ERROR generating fresh CSV files: {e}")
        raise


# ---------------------------------------------------------------------------
# Barcode label generation (from relabel_lib_plates_for_pooling.py)
# ---------------------------------------------------------------------------

def get_pooling_plates_data(db_path: str) -> List[Tuple[str, str]]:
    """
    Retrieve plate data for plates selected for pooling.

    Returns:
        List of (barcode, library_plate_container_barcode) tuples
        for plates where selected_for_pooling = 1, ordered by barcode.
    """
    if not os.path.exists(db_path):
        raise FileNotFoundError(f"Database file not found: {db_path}")

    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        query = """
        SELECT barcode, library_plate_container_barcode
        FROM individual_plates
        WHERE selected_for_pooling = 1
        ORDER BY barcode
        """

        cursor.execute(query)
        results = cursor.fetchall()
        conn.close()

        # Filter out rows where either barcode or container_barcode is None
        filtered_results = [
            (barcode, container_barcode)
            for barcode, container_barcode in results
            if barcode is not None and container_barcode is not None
        ]

        return filtered_results

    except sqlite3.Error as e:
        print(f"ERROR querying pooling plates data: {e}")
        print("SCRIPT TERMINATED: Cannot retrieve plate data from database")
        print("Ensure Script 1 (process_grid_tables_and_generate_barcodes.py) has been run first.")
        sys.exit()


def get_proposal_value(db_path: str) -> str:
    """
    Retrieve the unique proposal value from the sample_metadata table.
    """
    if not os.path.exists(db_path):
        raise FileNotFoundError(f"Database file not found: {db_path}")

    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()

        query = "SELECT DISTINCT Proposal FROM sample_metadata WHERE Proposal IS NOT NULL"
        cursor.execute(query)
        results = cursor.fetchall()
        conn.close()

        if not results:
            raise ValueError("No proposal values found in sample_metadata table")

        if len(results) > 1:
            proposals = [str(row[0]) for row in results]
            raise ValueError(f"Multiple proposal values found: {proposals}. Expected only one.")

        return str(results[0][0])

    except sqlite3.Error as e:
        raise sqlite3.Error(f"Database error: {e}")


def create_bartender_file(plate_data: List[Tuple[str, str]], output_dir: str, proposal: str) -> str:
    """
    Create a BarTender-compatible label file for library plate container barcodes.

    Args:
        plate_data: List of (barcode, library_plate_container_barcode) tuples
        output_dir: Directory where the BarTender file should be created
        proposal: Proposal value for filename

    Returns:
        Path to the created BarTender file
    """
    BARTENDER_HEADER = (
        '%BTW% /AF="\\\\BARTENDER\\shared\\templates\\ECHO_BCode8.btw" '
        '/D="%Trigger File Name%" /PRN="bcode8" /R=3 /P /DD\r\n\r\n%END%\r\n\r\n\r\n'
    )

    os.makedirs(output_dir, exist_ok=True)

    output_filename = f"BARTENDER_{proposal}_container_labels.txt"
    output_path = os.path.join(output_dir, output_filename)

    try:
        with open(output_path, 'w', newline='') as f:
            f.write(BARTENDER_HEADER)

            # Sort by container barcode for consistent ordering
            sorted_data = sorted(plate_data, key=lambda x: x[1])

            for i, (plate_barcode, container_barcode) in enumerate(sorted_data):
                f.write(f'{container_barcode},"{container_barcode}"\r\n')

                # Blank separator between entries (except after last)
                if i < len(sorted_data) - 1:
                    f.write(',\r\n')

            f.write('\r\n')

        return output_path

    except OSError as e:
        raise OSError(f"Failed to create BarTender file: {e}")


def find_blank_template(script_dir: Path) -> Path:
    """
    Find the BLANK_TEMPLATE_Pooling_Plate_Barcode_Scan_tool.xlsx file
    in the same directory as this script.
    """
    blank_template_path = script_dir / 'BLANK_TEMPLATE_Pooling_Plate_Barcode_Scan_tool.xlsx'

    if blank_template_path.exists():
        return blank_template_path
    else:
        raise FileNotFoundError(
            f"BLANK_TEMPLATE_Pooling_Plate_Barcode_Scan_tool.xlsx not found at: {blank_template_path}\n"
            f"Please ensure the template file is in the same directory as the script."
        )


def copy_template_file(template_path: str, output_dir: str, proposal: str) -> str:
    """
    Copy the blank template file to create a new working copy with proposal-based naming.

    Returns:
        Path to the copied file.
    """
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template file not found: {template_path}")

    os.makedirs(output_dir, exist_ok=True)

    output_filename = f"{proposal}_pool_label_scan_verificiation_tool.xlsx"
    output_path = os.path.join(output_dir, output_filename)

    try:
        shutil.copy2(template_path, output_path)
        return output_path
    except OSError as e:
        raise OSError(f"Failed to copy template file: {e}")


def populate_excel_template(excel_path: str, plate_data: List[Tuple[str, str]]) -> None:
    """
    Populate only the Expected Barcode column (Column C) of the Excel template.

    Template structure (confirmed from actual template inspection):
        Row 1:  Headers (sample, type, Expected Barcode, Checker, Scanned barcode)
        Row 2:  Sample 1 - "plate name"  → C2 = plate barcode (e.g. XUPVQ-1)
        Row 3:  Sample 1 - "LIMS ID"     → C3 = container barcode (e.g. 27-810254)
        Row 4:  Sample 2 - "plate name"  → C4 = plate barcode
        Row 5:  Sample 2 - "LIMS ID"     → C5 = container barcode
        ...
        Row 40: Sample 20 - "plate name"
        Row 41: Sample 20 - "LIMS ID"

    Row calculation:
        plate_name_row = 2 + (i * 2)
        lims_id_row    = 3 + (i * 2)

    ONLY Column C is modified. Columns A, B, D, E are left untouched.

    Args:
        excel_path: Path to the Excel file to populate (already copied from template)
        plate_data: List of (plate_barcode, container_barcode) tuples, ordered by plate
    """
    MAX_PLATES = 20  # Template handles 20 plates (rows 2-41, 2 rows each)

    if len(plate_data) > MAX_PLATES:
        raise ValueError(
            f"Too many plates ({len(plate_data)}) for template capacity ({MAX_PLATES})"
        )

    try:
        workbook = openpyxl.load_workbook(excel_path)
        worksheet = workbook['Sheet1']

        for i, (plate_barcode, container_barcode) in enumerate(plate_data):
            plate_name_row = 2 + (i * 2)   # Row for "plate name" type
            lims_id_row    = 3 + (i * 2)   # Row for "LIMS ID" type

            # Column C = column index 3
            worksheet.cell(row=plate_name_row, column=3, value=plate_barcode)
            worksheet.cell(row=lims_id_row,    column=3, value=container_barcode)

        workbook.save(excel_path)
        workbook.close()

    except Exception as e:
        raise Exception(f"Failed to populate Excel template: {e}")


# ---------------------------------------------------------------------------
# Main workflow
# ---------------------------------------------------------------------------

def main():
    """Main execution function for grid table processing and barcode generation."""
    parser = argparse.ArgumentParser(
        description=(
            "Process grid tables, update database, and generate barcode scanning materials. "
            "Run from within a project directory containing project_summary.db."
        )
    )
    parser.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="Enable verbose output",
    )
    args = parser.parse_args()

    try:
        print("=" * 60)
        print("Starting grid table processing and barcode generation...")
        print("=" * 60)

        base_dir = Path.cwd()
        script_dir = Path(__file__).parent

        if not base_dir.exists():
            raise FileNotFoundError(f"Base directory not found: {base_dir}")

        # ------------------------------------------------------------------
        # Step 1: Read database
        # ------------------------------------------------------------------
        print("\n[Step 1] Reading project database...")
        master_plate_df, individual_plates_df = read_project_database(base_dir)
        print(f"  ✓ master_plate_data: {len(master_plate_df)} rows")
        print(f"  ✓ individual_plates: {len(individual_plates_df)} rows")

        # ------------------------------------------------------------------
        # Step 2: Identify expected grid samples
        # ------------------------------------------------------------------
        print("\n[Step 2] Identifying samples selected for pooling...")
        expected_samples, selected_plates = identify_expected_grid_samples(
            master_plate_df, individual_plates_df
        )
        print(f"  ✓ {len(selected_plates)} plate(s) selected for pooling")
        print(f"  ✓ {len(expected_samples)} sample(s) expected in grid tables")

        # ------------------------------------------------------------------
        # Step 3: Find and validate grid table files
        # ------------------------------------------------------------------
        print("\n[Step 3] Finding grid table files...")
        grid_table_files = find_all_grid_tables(base_dir)
        print(f"  ✓ Found {len(grid_table_files)} valid grid table file(s):")
        for f in grid_table_files:
            print(f"    - {Path(f).name}")

        # ------------------------------------------------------------------
        # Step 4: Read and validate grid tables
        # ------------------------------------------------------------------
        print("\n[Step 4] Reading and validating grid tables...")
        grid_dataframes, combined_grid_df = read_multiple_grid_tables(grid_table_files)
        detect_duplicate_samples(grid_dataframes)
        validate_grid_table_completeness(expected_samples, combined_grid_df)
        print(f"  ✓ Grid tables valid: {len(combined_grid_df)} total rows, no duplicates")

        # ------------------------------------------------------------------
        # Step 5: Extract barcode mapping
        # ------------------------------------------------------------------
        print("\n[Step 5] Extracting Library Plate Container Barcode mapping...")
        barcode_mapping = extract_library_plate_container_barcode_mapping(combined_grid_df)
        for plate, container in barcode_mapping.items():
            print(f"  ✓ {plate} → {container}")

        # ------------------------------------------------------------------
        # Step 6: Merge grid data with master plate data
        # ------------------------------------------------------------------
        print("\n[Step 6] Merging grid data with master plate data...")
        merged_df = validate_and_merge_data(master_plate_df, expected_samples, combined_grid_df)
        print(f"  ✓ Merged dataframe: {len(merged_df)} rows")

        # ------------------------------------------------------------------
        # Step 7: Archive database and update tables
        # ------------------------------------------------------------------
        print("\n[Step 7] Archiving database and updating tables...")
        archive_db_path = archive_database_file(base_dir)
        if archive_db_path:
            print(f"  ✓ Database archived: {archive_db_path.name}")

        processed_plate_barcodes = list(
            merged_df[merged_df['Nucleic Acid ID'].notna()]['Plate_Barcode'].unique()
        )

        update_individual_plates_with_container_barcodes(
            base_dir, processed_plate_barcodes, barcode_mapping
        )
        print(f"  ✓ individual_plates updated for {len(processed_plate_barcodes)} plate(s)")

        update_master_plate_data_table(base_dir, merged_df)
        print(f"  ✓ master_plate_data table updated")

        # ------------------------------------------------------------------
        # Step 8: Archive grid files and regenerate CSVs
        # ------------------------------------------------------------------
        print("\n[Step 8] Archiving grid files and regenerating CSV files...")
        archive_grid_table_files(base_dir, grid_table_files)
        print(f"  ✓ {len(grid_table_files)} grid file(s) archived")

        archive_csv_files(base_dir)
        generate_fresh_csv_files(base_dir, merged_df)
        print(f"  ✓ Fresh CSV files generated")

        # ------------------------------------------------------------------
        # Step 9: Generate barcode scanning materials
        # ------------------------------------------------------------------
        print("\n[Step 9] Generating barcode scanning materials...")

        db_path = str(base_dir / "project_summary.db")
        proposal = get_proposal_value(db_path)
        print(f"  ✓ Proposal: {proposal}")

        # Get updated plate data (now includes library_plate_container_barcode)
        plate_data = get_pooling_plates_data(db_path)

        if not plate_data:
            print("  ⚠️  No plates found with selected_for_pooling = 1 and container barcodes")
            print("SCRIPT TERMINATED: Cannot generate barcode materials without plate data")
            sys.exit()

        print(f"  ✓ {len(plate_data)} plate(s) ready for barcode generation:")
        for plate_barcode, container_barcode in plate_data:
            print(f"    - {plate_barcode} → {container_barcode}")

        output_dir = str(
            base_dir / "4_plate_selection_and_pooling" / "B_new_plate_barcode_labels"
        )

        # Find blank template (in same directory as this script)
        template_path = find_blank_template(script_dir)

        # Copy template and populate Column C only
        excel_output_path = copy_template_file(str(template_path), output_dir, proposal)
        populate_excel_template(excel_output_path, plate_data)
        print(f"  ✓ Excel template created: {Path(excel_output_path).name}")

        # Create BarTender label file
        bartender_path = create_bartender_file(plate_data, output_dir, proposal)
        print(f"  ✓ BarTender file created: {Path(bartender_path).name}")

        # ------------------------------------------------------------------
        # Done
        # ------------------------------------------------------------------
        print("\n" + "=" * 60)
        print("🎉 Grid processing and barcode generation completed successfully!")
        print(f"📊 Processed {len(processed_plate_barcodes)} plate(s)")
        print(f"📁 Output directory: 4_plate_selection_and_pooling/B_new_plate_barcode_labels/")
        print("\nNEXT STEP: Open the Excel file, complete barcode scanning,")
        print("           then run verify_scanning_and_generate_ESP_files.py")
        print("=" * 60)

        create_success_marker()

    except Exception as e:
        print(f"\nERROR: Grid processing failed: {e}")
        sys.exit()


if __name__ == "__main__":
    main()
