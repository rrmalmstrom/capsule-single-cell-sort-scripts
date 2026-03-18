#!/usr/bin/env python3
"""
Script to create a populated barcode scan template for capsule SPS pooling.

This script:
1. Copies the blank template Excel file
2. Retrieves data from the individual_plates table in project_summary.db
3. Populates the template with plate barcodes and container barcodes
4. Saves the populated file with a timestamped name

Author: Generated for capsule sort scripts
Date: 2026-03-16
"""

import sqlite3
import shutil
import openpyxl
import os
import sys
import argparse
from datetime import datetime
from pathlib import Path
from typing import List, Tuple, Optional


def create_success_marker():
    """Create success marker file for workflow manager integration."""
    script_name = Path(__file__).stem
    status_dir = Path(".workflow_status")
    status_dir.mkdir(exist_ok=True)
    success_file = status_dir / f"{script_name}.success"

    try:
        with open(success_file, "w") as f:
            f.write(f"SUCCESS: {script_name} completed at {datetime.now()}\n")
        print(f"✅ Success marker created: {success_file}")
    except Exception as e:
        print(f"FATAL ERROR: Could not create success marker: {e}")
        print("Laboratory automation requires workflow integration for safety.")
        sys.exit()


def get_pooling_plates_data(db_path: str) -> List[Tuple[str, str]]:
    """
    Retrieve plate data for plates selected for pooling.
    
    Args:
        db_path: Path to the project_summary.db file
        
    Returns:
        List of tuples containing (barcode, library_plate_container_barcode)
        for plates where selected_for_pooling = 1
        
    Raises:
        sqlite3.Error: If database connection or query fails
        FileNotFoundError: If database file doesn't exist
    """
    if not os.path.exists(db_path):
        raise FileNotFoundError(f"Database file not found: {db_path}")
    
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Query for plates selected for pooling
        query = """
        SELECT barcode, library_plate_container_barcode
        FROM individual_plates
        WHERE selected_for_pooling = 1
        ORDER BY barcode
        """
        
        cursor.execute(query)
        results = cursor.fetchall()
        
        # Filter out rows where either barcode or container_barcode is None
        filtered_results = [
            (barcode, container_barcode) 
            for barcode, container_barcode in results 
            if barcode is not None and container_barcode is not None
        ]
        
        conn.close()
        return filtered_results
        
    except sqlite3.Error as e:
        raise sqlite3.Error(f"Database error: {e}")


def get_proposal_value(db_path: str) -> str:
    """
    Retrieve the unique proposal value from the sample_metadata table.
    
    Args:
        db_path: Path to the project_summary.db file
        
    Returns:
        The proposal value as a string
        
    Raises:
        sqlite3.Error: If database connection or query fails
        ValueError: If no proposal found or multiple proposals found
    """
    if not os.path.exists(db_path):
        raise FileNotFoundError(f"Database file not found: {db_path}")
    
    try:
        conn = sqlite3.connect(db_path)
        cursor = conn.cursor()
        
        # Query for unique proposal values
        query = "SELECT DISTINCT Proposal FROM sample_metadata WHERE Proposal IS NOT NULL"
        cursor.execute(query)
        results = cursor.fetchall()
        
        conn.close()
        
        if not results:
            raise ValueError("No proposal values found in sample_metadata table")
        
        if len(results) > 1:
            proposals = [str(row[0]) for row in results]
            raise ValueError(f"Multiple proposal values found: {proposals}. Expected only one.")
        
        return str(results[0][0])
        
    except sqlite3.Error as e:
        raise sqlite3.Error(f"Database error: {e}")


def create_bartender_file(plate_data: List[Tuple[str, str]], output_dir: str, proposal: str) -> str:
    """
    Create a BarTender-compatible label file for library plate container barcodes.
    
    Args:
        plate_data: List of (barcode, library_plate_container_barcode) tuples
        output_dir: Directory where the BarTender file should be created
        proposal: Proposal value for filename
        
    Returns:
        Path to the created BarTender file
        
    Raises:
        OSError: If file creation fails
    """
    # BarTender header from the reference script
    BARTENDER_HEADER = '%BTW% /AF="\\\\BARTENDER\\shared\\templates\\ECHO_BCode8.btw" /D="%Trigger File Name%" /PRN="bcode8" /R=3 /P /DD\r\n\r\n%END%\r\n\r\n\r\n'
    
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Generate filename with proposal value
    output_filename = f"BARTENDER_{proposal}_container_labels.txt"
    output_path = os.path.join(output_dir, output_filename)
    
    try:
        with open(output_path, 'w', newline='') as f:
            # Write header
            f.write(BARTENDER_HEADER)
            
            # Sort by container barcode for consistent ordering
            sorted_data = sorted(plate_data, key=lambda x: x[1])  # Sort by container barcode
            
            # Write container barcodes in the BarTender format
            for i, (plate_barcode, container_barcode) in enumerate(sorted_data):
                # Use container barcode as both the barcode and label
                f.write(f'{container_barcode},"{container_barcode}"\r\n')
                
                # Add blank separator line between entries (except after last entry)
                if i < len(sorted_data) - 1:
                    f.write(',\r\n')
            
            # Add proper trailing empty lines for BarTender compatibility
            f.write('\r\n')
        
        return output_path
        
    except OSError as e:
        raise OSError(f"Failed to create BarTender file: {e}")


def copy_template_file(template_path: str, output_dir: str, proposal: str) -> str:
    """
    Copy the blank template file to create a new working copy with proposal-based naming.
    
    Args:
        template_path: Path to the blank template Excel file
        output_dir: Directory where the copy should be created
        proposal: Proposal value to use in filename
        
    Returns:
        Path to the copied file with proposal-based name
        
    Raises:
        FileNotFoundError: If template file doesn't exist
        OSError: If file copy operation fails
    """
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template file not found: {template_path}")
    
    # Create output directory if it doesn't exist
    os.makedirs(output_dir, exist_ok=True)
    
    # Generate filename with proposal value
    output_filename = f"{proposal}_pool_label_scan_verificiation_tool.xlsx"
    output_path = os.path.join(output_dir, output_filename)
    
    try:
        shutil.copy2(template_path, output_path)
        return output_path
    except OSError as e:
        raise OSError(f"Failed to copy template file: {e}")


def populate_excel_template(excel_path: str, plate_data: List[Tuple[str, str]]) -> None:
    """
    Populate the Excel template with plate barcode data.
    
    Args:
        excel_path: Path to the Excel file to populate
        plate_data: List of (barcode, container_barcode) tuples
        
    Raises:
        openpyxl.utils.exceptions.InvalidFileException: If Excel file is invalid
        ValueError: If too many plates for template capacity
    """
    # Template can handle up to 20 plates (rows 3-42, alternating M/N columns)
    MAX_PLATES = 20
    
    if len(plate_data) > MAX_PLATES:
        raise ValueError(f"Too many plates ({len(plate_data)}) for template capacity ({MAX_PLATES})")
    
    try:
        workbook = openpyxl.load_workbook(excel_path)
        worksheet = workbook['Sheet1']
        
        # Populate data starting from row 3
        for i, (barcode, container_barcode) in enumerate(plate_data):
            row = 3 + i
            
            # Column M: barcode (plate identifier)
            worksheet.cell(row=row, column=13, value=barcode)
            
            # Column N: library_plate_container_barcode
            worksheet.cell(row=row, column=14, value=container_barcode)
        
        # Save the workbook
        workbook.save(excel_path)
        workbook.close()
        
    except Exception as e:
        raise Exception(f"Failed to populate Excel template: {e}")


def validate_inputs(db_path: str, template_path: str) -> None:
    """
    Validate that required input files exist and are accessible.
    
    Args:
        db_path: Path to the database file
        template_path: Path to the template Excel file
        
    Raises:
        FileNotFoundError: If any required file is missing
    """
    if not os.path.exists(db_path):
        raise FileNotFoundError(f"Database file not found: {db_path}")
    
    if not os.path.exists(template_path):
        raise FileNotFoundError(f"Template file not found: {template_path}")


def find_blank_template():
    """Find the BLANK_TEMPLATE_Pooling_Plate_Barcode_Scan_tool.xlsx file in the scripts directory."""
    # Look for the file in the same directory as this script
    blank_template_path = SCRIPT_DIR / 'BLANK_TEMPLATE_Pooling_Plate_Barcode_Scan_tool.xlsx'
    
    if blank_template_path.exists():
        return blank_template_path
    else:
        raise FileNotFoundError(
            f"BLANK_TEMPLATE_Pooling_Plate_Barcode_Scan_tool.xlsx not found at: {blank_template_path}\n"
            f"Please ensure the template file is in the same directory as the script."
        )


def create_barcode_scan_template() -> str:
    """
    Main function to create a populated barcode scan template.
    
    This function should be run from within a project directory.
    It will:
    1. Look for project_summary.db in the current directory
    2. Create output in 4_plate_selection_and_pooling/C_pooling_barcode_labels/
    3. Use the template from the script's directory
    
    Returns:
        Path to the created Excel file
        
    Raises:
        Various exceptions from called functions
    """
    # Database path in current project directory
    db_path = PROJECT_DIR / "project_summary.db"
    
    # Output directory
    output_dir = PROJECT_DIR / "4_plate_selection_and_pooling" / "C_pooling_barcode_labels"
    
    # Check if the template file exists before proceeding
    try:
        template_path = find_blank_template()
    except FileNotFoundError as e:
        print(f"ERROR: {e}")
        raise
    
    # Validate inputs (convert Path objects to strings for the validation function)
    validate_inputs(str(db_path), str(template_path))
    
    # Get proposal value from sample_metadata table
    proposal = get_proposal_value(str(db_path))
    
    # Get plate data from database
    plate_data = get_pooling_plates_data(str(db_path))
    
    if not plate_data:
        print("Warning: No plates found with selected_for_pooling = 1")
        return None
    
    print(f"Processing {len(plate_data)} plates for proposal {proposal}")
    
    # Copy template file with proposal-based naming
    output_path = copy_template_file(str(template_path), str(output_dir), proposal)
    
    # Populate the template
    populate_excel_template(output_path, plate_data)
    
    # Create BarTender file for container labels
    bartender_path = create_bartender_file(plate_data, str(output_dir), proposal)
    
    print(f"✓ Created: {os.path.basename(output_path)}")
    print(f"✓ Created: {os.path.basename(bartender_path)}")
    return output_path


def main():
    """Command-line interface for the script."""
    global PROJECT_DIR, SCRIPT_DIR
    PROJECT_DIR = Path.cwd()
    SCRIPT_DIR = Path(__file__).parent  # Directory where this script is located
    
    parser = argparse.ArgumentParser(
        description="Create populated barcode scan template for capsule SPS pooling. "
                   "Run this script from within a project directory containing project_summary.db"
    )
    parser.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="Enable verbose output"
    )
    
    args = parser.parse_args()
    
    try:
        output_path = create_barcode_scan_template()
        if not output_path:
            print("⚠ No plates selected for pooling")
            return 1

    except Exception as e:
        print(f"✗ Error: {e}")
        return 1

    # Create success marker for workflow manager
    create_success_marker()

    return 0


if __name__ == "__main__":
    exit(main())